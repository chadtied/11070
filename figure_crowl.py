from seleniumbase import Driver
from selenium.webdriver.common.by import By
import openpyxl
from openpyxl.styles import Font



workbook= openpyxl.Workbook()
workpage= workbook.create_sheet("List1",0)
imfor= dict()
driver= Driver(uc= True, incognito= False)
red_font= Font(color= 'FF0000')
green_font= Font(color= '00FF00')

try:
    driver.get("https://www.wantgoo.com/global")
    figures= driver.find_elements(By.CSS_SELECTOR,"#mainIndex > div > div > table > tbody> tr")
except:
    print("搜集指數資料錯誤")
    
for figure in figures:
    try:
        if figure.find_element(By.CSS_SELECTOR, "td.lt").text== "費城半導體":
            imfor['SOX']= {}
            imfor['SOX']['CLOSE']= float(figure.find_element(By.NAME, "close").text)
            try:imfor['SOX']['CHANGE']= float((figure.find_element(By.NAME, "change").text)[1:])
            except:imfor['SOX']['CHANGE']= 0
            imfor['SOX']['RATE']= float(figure.find_element(By.NAME, "changeRate").text)
    except:
        print("費城半導體資料採掘錯誤")
    
    try:
        if figure.find_element(By.CSS_SELECTOR, "td.lt").text== "加權指數":
            imfor['TAIEX']= {}
            imfor['TAIEX']['CLOSE']= float(figure.find_element(By.NAME, "close").text)
            try:imfor['TAIEX']['CHANGE']= float((figure.find_element(By.NAME, "change").text)[1:])
            except:imfor['TAIEX']['CHANGE']= 0
            imfor['TAIEX']['RATE']= float(figure.find_element(By.NAME, "changeRate").text)
    except:
        print("加權指數資料採掘錯誤")
        
        
    try:
        if figure.find_element(By.CSS_SELECTOR, "td.lt").text== "台指期":
            imfor['TIF']= {}
            imfor['TIF']['CLOSE']= float(figure.find_element(By.NAME, "close").text)
            try:imfor['TIF']['CHANGE']= float((figure.find_element(By.NAME, "change").text)[1:])
            except:imfor['TIF']['CHANGE']= 0
            imfor['TIF']['RATE']= float(figure.find_element(By.NAME, "changeRate").text)
    except:
        print("台指期資料採掘錯誤")
    
    try:
        if figure.find_element(By.CSS_SELECTOR, "td.lt").text== "台積電ADR":
            imfor['TSMC']= {}
            imfor['TSMC']['CLOSE']= float(figure.find_element(By.NAME, "close").text)
            try:imfor['TSMC']['CHANGE']= float((figure.find_element(By.NAME, "change").text)[1:])
            except:imfor['TSMC']['CHANGE']= 0
            imfor['TSMC']['RATE']= float(figure.find_element(By.NAME, "changeRate").text)
    except:
        print("台積電ADR資料採掘錯誤")
    
    try:
        if figure.find_element(By.CSS_SELECTOR, "td.lt").text== "道瓊指數":
            imfor['DJIA']= {}
            imfor['DJIA']['CLOSE']= float(figure.find_element(By.NAME, "close").text)
            try:imfor['DJIA']['CHANGE']= float((figure.find_element(By.NAME, "change").text)[1:])
            except:imfor['DJIA']['CHANGE']= 0
            imfor['DJIA']['RATE']= float(figure.find_element(By.NAME, "changeRate").text)
    except:
        print("道瓊指數資料採掘錯誤")

print(imfor)

workpage.cell(1,1).value= "指標名稱"
workpage.cell(1,2).value= "最近收盤價"
workpage.cell(1,3).value= "漲跌點數"
workpage.cell(1,4).value= "漲跌幅"
row= 2
for key, fig_dict in imfor.items():
 workpage.cell(row,1).value= key
 for fig_key,  fig_data in fig_dict.items():
    if fig_key== 'CLOSE': 
        workpage.cell(row,2).value= fig_data
        if(fig_data>= 0):  workpage.cell(row,2).font= red_font
        else: workpage.cell(row,2).font= green_font
    elif fig_key== 'CHANGE':
        workpage.cell(row,3).value= fig_data
        if(fig_data>= 0):  workpage.cell(row,3).font= red_font
        else: workpage.cell(row,2).font= green_font
    elif fig_key== 'RATE': 
        workpage.cell(row,4).value= fig_data
        if(fig_data>= 0):  workpage.cell(row,4).font= red_font
        else: workpage.cell(row,2).font= green_font
 row+= 1

 workbook.save("./new.xlsx")
